from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from acms_cop.classifiers.behavioral_state_classifier import classify_behavioral_state
from acms_cop.classifiers.causal_status_classifier import classify_causal_status
from acms_cop.classifiers.flow_collision_classifier import classify_flow_collision
from acms_cop.common import first_dict, json_dumps, safe_float, safe_int


HEADER_ALIASES = {
    "ticker": {"ticker", "symbol"},
    "flow_bias": {"bias", "flowbias", "flow_bias"},
    "main_net_flow": {"mainnet", "main_net", "mainnetflow", "main_net_flow"},
    "super_large_net_flow": {"superlargenet", "super_large_net", "superlargenetflow"},
    "large_net_flow": {"largenet", "large_net", "largenetflow"},
    "medium_net_flow": {"mediumnet", "medium_net", "mediumnetflow"},
    "small_net_flow": {"smallnet", "small_net", "smallnetflow"},
    "in_flow": {"inflow", "in_flow"},
    "snapshot": {"snapshot"},
    "cycle_ts": {"cyclets", "cycle_ts", "cyclets"},
}


def _norm(value: Any) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def _header_map(headers: List[Any]) -> Dict[str, int]:
    normalized = [_norm(h) for h in headers]
    out: Dict[str, int] = {}
    for target, aliases in HEADER_ALIASES.items():
        for idx, name in enumerate(normalized):
            if name in aliases:
                out[target] = idx
                break
    return out


def _price_row(dataset: Dict[str, Any], ticker: str) -> Dict[str, Any]:
    live = first_dict(dataset.get("live_prices"))
    return first_dict(live.get(ticker))


def _metadata_row(dataset: Dict[str, Any], ticker: str) -> Dict[str, Any]:
    return first_dict(
        first_dict(dataset.get("security_master")).get(ticker),
        first_dict(dataset.get("fundamentals")).get(ticker),
        first_dict(dataset.get("analyst_targets")).get(ticker),
    )


def _direction(day_return: float | None) -> str:
    if day_return is None:
        return "FLAT"
    if day_return > 0.05:
        return "UP"
    if day_return < -0.05:
        return "DOWN"
    return "FLAT"


def _row_from_cells(headers: List[Any], values: List[Any]) -> Dict[str, Any]:
    hm = _header_map(headers)
    row = {key: values[idx] if idx < len(values) else None for key, idx in hm.items()}
    return row


def extract_ticker_cycles(dataset: Dict[str, Any], workbook_path: str | Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for ACMS ticker extraction") from exc

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = next((name for name in wb.sheetnames if name.lower().strip() == "capital flow"), None)
    if not sheet_name:
        raise ValueError("Capital Flow sheet not found in V3 workbook")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    extracted: List[Dict[str, Any]] = []
    regime = first_dict(dataset.get("regime"))
    cross = first_dict(dataset.get("cross_market_confirmation"))
    liquidity_stress = bool(first_dict(cross.get("credit_liquidity_stress")).get("stress_active"))

    for values in rows[1:]:
        raw = _row_from_cells(headers, list(values))
        ticker = str(raw.get("ticker") or "").strip().upper()
        if not ticker:
            continue
        price_info = _price_row(dataset, ticker)
        meta = _metadata_row(dataset, ticker)
        day_return = safe_float(price_info.get("chg_pct") or price_info.get("change_pct") or price_info.get("day_return"))
        price_direction = _direction(day_return)
        flow_bias = str(raw.get("flow_bias") or "").strip().upper()
        causal = classify_causal_status(source_count=0, direction=price_direction)
        collision = classify_flow_collision(price_direction, flow_bias, raw.get("main_net_flow"), None, causal)
        behavior = classify_behavioral_state(
            price_direction,
            flow_bias,
            causal_status=causal,
            regime_label=regime.get("regime") or regime.get("regime_short"),
            liquidity_stress_active=liquidity_stress,
            vix_rising=False,
            credit_weakening=liquidity_stress,
        )
        row = {
            "ticker": ticker,
            "company_name": meta.get("company_name") or meta.get("name") or meta.get("company"),
            "theme": meta.get("theme") or meta.get("sector") or meta.get("industry"),
            "sleeve": meta.get("sleeve") or meta.get("asset_type"),
            "price": safe_float(price_info.get("price") or price_info.get("last") or price_info.get("last_price")),
            "day_return": day_return,
            "volume": safe_int(price_info.get("volume")),
            "volume_zscore": safe_float(price_info.get("volume_zscore")),
            "flow_bias": flow_bias,
            "main_net_flow": safe_float(raw.get("main_net_flow")),
            "super_large_net_flow": safe_float(raw.get("super_large_net_flow")),
            "large_net_flow": safe_float(raw.get("large_net_flow")),
            "medium_net_flow": safe_float(raw.get("medium_net_flow")),
            "small_net_flow": safe_float(raw.get("small_net_flow")),
            "in_flow": safe_float(raw.get("in_flow")),
            "options_bias": price_info.get("options_bias"),
            "put_call_ratio": safe_float(price_info.get("put_call_ratio")),
            "hedge_demand_score": safe_float(price_info.get("hedge_demand_score")),
            "price_direction": collision["price_direction"],
            "flow_direction": collision["flow_direction"],
            "causal_status": causal,
            "acms_state": behavior["acms_state"],
            "secondary_state": behavior["secondary_state"],
            "acms_confidence": behavior["confidence"],
            "cio_meaning": behavior["cio_meaning"],
            "recommended_posture": behavior["recommended_posture"],
            "blocked_actions_json": json_dumps(behavior["blocked_actions_json"]),
            "data_quality_flag": "OK",
            "raw_fields_json": json_dumps(raw),
        }
        extracted.append(row)
    return extracted

